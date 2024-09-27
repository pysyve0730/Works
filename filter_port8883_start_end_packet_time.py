import re

import openpyxl
from openpyxl import Workbook
#packet_number connot repeat, so delete repeat data
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
f = open("X5A_0401_A010_new.txt", "r",encoding='UTF-8') #截掉第一筆DNS的.txt


lines = f.readlines()
tmp = []
hand_shake_type = []
info = []
start_pkg_num = []
start_pkg_time = []
end_pkg_num = []
end_pkg_time = []


source_pkg_num = 6
source_pkg_time = 1
dist_pkg_num = 8
dist_pkg_time = 1

for line in lines:
    #print(line) #Print all information
    tmp = re.split(r'\[', line)

    if(len(tmp)==2):
        info = tmp[0].split()
        hand_shake_type = tmp[1].split()
        #print(info)
        if len(info)==9:
            
            if "8883" in info and "SYN]" in hand_shake_type[0]:
                start_pkg_num.append(info[source_pkg_num])
                start_pkg_time.append(info[source_pkg_time])

            elif "8883" in info and ("FIN," in hand_shake_type[0] or "RST," in hand_shake_type[0]):
                end_pkg_num.append(info[dist_pkg_num])
                end_pkg_time.append(info[dist_pkg_time])

start_dict = dict(zip(start_pkg_num, start_pkg_time))
#print("Start has "+str(len(start_dict))+" rows:"+str(start_dict))

print("=======================================")

end_dict = dict(zip(end_pkg_num, end_pkg_time))
#print("End has "+str(len(end_dict))+" rows:"+str(end_dict))


excel_file  = '0401_24hours_X5A_2.0.0-A010_direct_write.xlsx'

wb = openpyxl.load_workbook(excel_file, data_only=True)
s1 = wb['chang-24hours']            # 開啟工作表
ws = wb.active

start_row = 4
for i, (start_key, start_value) in enumerate(start_dict.items(), start=0):
        ws.cell(row=start_row + i, column=7, value=start_key)
        ws.cell(row=start_row + i, column=8, value=start_value)
        #填充顏色
        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        ws.cell(row=start_row+i, column=7).fill = fill
        ws.cell(row=start_row+i, column=8).fill = fill
        #加入外框
        border = Border(left=Side(border_style='thin'), 
                    right=Side(border_style='thin'), 
                    top=Side(border_style='thin'), 
                    bottom=Side(border_style='thin'))
        ws.cell(row=start_row+i, column=7).border = border
        ws.cell(row=start_row+i, column=8).border = border
        #選擇字體
        font = Font(name='Calibri', size=11, bold=False, italic=False, color='000000')
        
        ws.cell(row=start_row+i, column=7).font = font
        ws.cell(row=start_row+i, column=8).font = font


for i, (end_key, end_value) in enumerate(end_dict.items(), start=0):     
        ws.cell(row=start_row+i, column=10, value=end_key)
        ws.cell(row=start_row+i, column=11, value=end_value)
        
        #填充顏色
        fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        ws.cell(row=start_row+i, column=10).fill = fill
        ws.cell(row=start_row+i, column=11).fill = fill
        #加入外框
        border = Border(left=Side(border_style='thin'), 
                    right=Side(border_style='thin'), 
                    top=Side(border_style='thin'), 
                    bottom=Side(border_style='thin'))
        ws.cell(row=start_row+i, column=10).border = border
        ws.cell(row=start_row+i, column=11).border = border
        
        
        font = Font(name='Calibri', size=11, bold=False, italic=False, color='000000')
        
        ws.cell(row=start_row+i, column=10).font = font
        ws.cell(row=start_row+i, column=11).font = font

#寫入delta time(second,minute)
det_sec = []
det_min = []
s_keys = list(start_dict.keys())
for i in range(1, len(s_keys)):
    s_k = s_keys[i]
    s_prev_k = s_keys[i-1]
    s_value = start_dict[s_k]
    det_sec.append(float(start_dict[s_k])- float(start_dict[s_prev_k]))
    det_min.append((float(start_dict[s_k])- float(start_dict[s_prev_k]))/60)
 
det = dict(zip(det_sec, det_min))        
for kk, (sec_val,min_val) in enumerate(det.items(), start=0):
    ws.cell(row=start_row+kk+1, column=3, value=sec_val)
    ws.cell(row=start_row+kk+1, column=4, value=min_val)


#寫入delta time(PCAP)
det_pcap=[]
for end_p, start_p in zip(end_dict.values(), start_dict.values()):
    det_pcap.append(float(end_p) - float(start_p))
for pp, pcap_val in enumerate(det_pcap, start=0):
    ws.cell(row=start_row+pp, column=13, value=pcap_val)
    

wb.save(excel_file)

#寫入start end Package number,time
# wb_start = Workbook()
# ws_start  = wb_start.active
# for start_pkg_num, start_time in start_dict.items():
    # ws_start.append([start_pkg_num, start_time])
# wb_start.save("output_0329_start.xlsx")


# wb_end = Workbook()
# ws_end = wb_end.active
# for end_pkg_num, end_pkg_time in end_dict.items():
    # ws_end.append([end_pkg_num, end_pkg_time])
# wb_end.save("output_0329_end.xlsx")















    

    

        
    









    











